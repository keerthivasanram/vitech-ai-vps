"""The package as files: a project folder, and the same folder as one download.

    Project/
        Review.md              <- read first
        Specification.pdf
        Drawing_GA.pdf
        Drawing_GA.svg
        Quotation.pdf
        BOM.xlsx
        Assumptions.md
        Project_Summary.md
        Traceability.md
        Cross_Reference.md
        package.json           <- the machine-readable manifest

Every renderer here is one that already exists: the specification and quotation
PDFs are the letterhead renderers the single-document endpoints use, and the
drawing PDF/SVG come from the same `Canvas` the studio approved. An export is
never a second rendering path, so a downloaded document cannot differ from the
one reviewed on screen.

A document that could not be produced is simply absent from the folder AND
recorded in the manifest with its reason. Writing an empty PDF so the folder
looks complete would be the worst possible outcome.
"""
import io
import json
import os
import re
import zipfile
from typing import Any, Optional

# openpyxl is already a declared dependency (the RAG loader reads XLSX with it),
# so the BOM spreadsheet adds no new package to the install.
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

PACKAGE_DIR = os.getenv("PACKAGE_DIR", os.path.join("data", "packages"))


def folder_name(pkg: dict) -> str:
    """A filesystem-safe project folder name derived from the package itself."""
    parts = [pkg.get("project"), pkg.get("client"), pkg.get("equipment")]
    label = next((str(p) for p in parts
                  if p and str(p) != "(to be completed)"), "Project")
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", label).strip("_") or "Project"
    rev = re.sub(r"[^A-Za-z0-9]+", "", str(pkg.get("revision") or "0"))
    return f"{slug}_Rev{rev}"


def _bom_xlsx(bom: dict, xref: Optional[dict] = None) -> bytes:
    """The bill of materials as a spreadsheet, because that is how a buyer uses it.

    Carries the package item id alongside each line, so a BOM row opened on its
    own can still be traced back to the drawing balloon and the specification.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    headers = ["Item", "Section", "Description", "Specification", "Qty", "Unit",
               "Weight (kg)", "Amount (INR)", "Basis", "Source"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    by_item = {}
    for entry in (xref or {}).get("items") or []:
        for name in entry.get("bom_items") or []:
            by_item.setdefault(str(name), entry["item_id"])

    for line in bom.get("lines") or []:
        ws.append([
            by_item.get(str(line.get("item")), ""),
            line.get("section"), line.get("item"), line.get("spec"),
            line.get("qty"), line.get("unit"), line.get("weight_kg"),
            line.get("amount"), line.get("basis"), line.get("source"),
        ])

    totals = bom.get("totals") or {}
    ws.append([])
    ws.append(["", "TOTAL (partial)", "", "", "", "",
               totals.get("weight_kg"), totals.get("amount")])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    # The uncosted list is the point of this document, not a footnote: it is
    # what the client still has to price.
    uncosted = bom.get("uncosted") or []
    if uncosted:
        ws.append([])
        ws.append(["UNCOSTED — no rate on file"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Item", "Reason"])
        for u in uncosted:
            if isinstance(u, dict):
                ws.append([u.get("item"), u.get("reason") or u.get("basis")])
            else:
                ws.append([str(u), ""])

    for note in bom.get("notes") or []:
        ws.append([])
        ws.append([str(note)])

    for i, width in enumerate([9, 20, 34, 44, 8, 8, 12, 14, 26, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_files(pkg: dict) -> tuple[dict[str, bytes], list[dict]]:
    """Render every document. Returns (filename -> bytes, per-file manifest).

    Failures are contained per document: one renderer raising must not cost the
    engineer the other eight files, so the failure is recorded against that file
    and the rest of the package still builds.
    """
    files: dict[str, bytes] = {}
    manifest: list[dict] = []

    def add(name: str, data: Optional[bytes], document: str, note: str = "") -> None:
        if data:
            files[name] = data
        manifest.append({"file": name, "document": document,
                         "written": bool(data), "note": note})

    md = pkg.get("markdown") or {}

    # Review first — it is the document the folder exists to deliver.
    add("Review.md", _utf8(md.get("review")), "review")
    add("Project_Summary.md", _utf8(md.get("summary")), "dashboard")
    add("Customer_Requirement.md", _utf8(md.get("requirement")), "requirement_summary")
    add("Assumptions.md", _utf8(md.get("assumptions")), "assumptions")
    add("Traceability.md", _utf8(md.get("traceability")), "traceability")
    add("Cross_Reference.md", _utf8(md.get("cross_reference")), "cross_reference")

    # --- specification PDF -------------------------------------------------
    spec = pkg.get("specification") or {}
    try:
        from ..specification_pdf import render_specification_pdf
        data = render_specification_pdf({
            "category_label": pkg.get("equipment"),
            "technical_details": spec.get("rows") or [],
            "confidence_pct": spec.get("confidence_pct"),
            "confidence_label": spec.get("confidence_label"),
            "text": spec.get("markdown") or "",
        }) if spec.get("rows") else None
        add("Specification.pdf", data, "specification",
            "" if data else "No engineering rows to print.")
    except Exception as exc:                                   # pragma: no cover
        add("Specification.pdf", None, "specification", f"Renderer failed: {exc}")

    # --- drawing: SVG + true-size PDF, both off the SAME canvas -------------
    drawing = pkg.get("drawing") or {}
    svg = drawing.get("svg")
    add("Drawing_GA.svg", _utf8(svg), "drawing",
        "" if svg else "No drawing was generated.")
    try:
        data = _drawing_pdf(pkg) if svg else None
        add("Drawing_GA.pdf", data, "drawing",
            "" if data else "No drawing was generated.")
    except Exception as exc:                                   # pragma: no cover
        add("Drawing_GA.pdf", None, "drawing", f"Renderer failed: {exc}")

    # --- quotation ---------------------------------------------------------
    quote = pkg.get("quotation") or {}
    try:
        from ..quotation_pdf import render_quotation_pdf
        data = render_quotation_pdf(quote) if quote else None
        add("Quotation.pdf", data, "quotation",
            "" if data else "No priced history, so no budgetary quotation.")
    except Exception as exc:                                   # pragma: no cover
        add("Quotation.pdf", None, "quotation", f"Renderer failed: {exc}")

    # --- BOM ---------------------------------------------------------------
    bom = pkg.get("bom") or {}
    try:
        data = _bom_xlsx(bom, pkg.get("cross_reference")) if bom.get("lines") else None
        add("BOM.xlsx", data, "bom", "" if data else "No BOM lines were derived.")
    except Exception as exc:                                   # pragma: no cover
        add("BOM.xlsx", None, "bom", f"Renderer failed: {exc}")
    if bom.get("bom_markdown"):
        add("BOM.md", _utf8(bom["bom_markdown"]), "bom")

    add("package.json", json.dumps(_manifest_doc(pkg, manifest), indent=2).encode(),
        "manifest")
    return files, manifest


def _drawing_pdf(pkg: dict) -> Optional[bytes]:
    """Re-compose the sheet and export it, so the PDF matches the reviewed SVG.

    The drawing package carries the SVG but not the Canvas the exporters need,
    and re-rendering from the stored `source` is exactly what the studio's own
    export does — one composition path, never a second one that could drift.
    """
    from ..drawing.drawing_service import compose
    from ..drawing import export as exporter

    source = (pkg.get("drawing") or {}).get("_source")
    if not source:
        return None
    canvas, _ = compose(source["spec"], **source.get("options") or {})
    return exporter.to_pdf(canvas)


def _manifest_doc(pkg: dict, files: list[dict]) -> dict[str, Any]:
    return {
        "ref": pkg.get("ref"),
        "project": pkg.get("project"),
        "client": pkg.get("client"),
        "equipment": pkg.get("equipment"),
        "revision": pkg.get("revision"),
        "generated": pkg.get("generated"),
        "read_first": "Review.md",
        "release_status": (pkg.get("review") or {}).get("release_status"),
        "verdict": (pkg.get("review") or {}).get("verdict"),
        "documents": pkg.get("manifest"),
        "files": files,
        "dashboard": pkg.get("dashboard"),
    }


def write_package(pkg: dict, root: Optional[str] = None) -> dict[str, Any]:
    """Write the project folder to disk and return what was written."""
    files, manifest = build_files(pkg)
    base = os.path.join(root or PACKAGE_DIR, folder_name(pkg))
    os.makedirs(base, exist_ok=True)
    for name, data in files.items():
        with open(os.path.join(base, name), "wb") as fh:
            fh.write(data)
    return {"ok": True, "folder": base, "files": manifest,
            "written": sorted(files)}


def zip_package(pkg: dict) -> tuple[bytes, str]:
    """The whole project folder as a single download."""
    files, _ = build_files(pkg)
    folder = folder_name(pkg)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in sorted(files):
            zf.writestr(f"{folder}/{name}", files[name])
    return buf.getvalue(), f"{folder}.zip"


def _utf8(text) -> Optional[bytes]:
    return str(text).encode("utf-8") if text else None
